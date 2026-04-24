"""
Writes a list of row dicts (from parser.py) to an .xlsx file matching the
PPVPOC reference format:

  A: Promo Number (text, e.g. "1", "1-1")
  B: Name          (original capitalisation)
  C: Promo Number  (duplicate of A)
  D: Promo Name    (duplicate of B)
  E: Cue           (uppercased body text)
  F: Notes         (empty)
  G: =LEN(E{row})  (formula)

Row 1 is the header row; data starts at row 2.
Numbers are written as plain text to prevent Excel/Sheets from
interpreting values like "1-1" as dates.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADERS = ["", "Name", "Promo Number", "Promo Name", "Cue", "Notes", "Character"]

# Column widths (approximate, in character units)
COL_WIDTHS = {
    "A": 12,
    "B": 40,
    "C": 14,
    "D": 40,
    "E": 80,
    "F": 20,
    "G": 12,
}


def write_xlsx(rows: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row
    header_font = Font(bold=True)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        number = row["number"]
        name = row["name"]
        cue = row["cue"]

        # A: Promo Number as text (prevents date misinterpretation of "1-1")
        a = ws.cell(row=row_idx, column=1, value=number)
        a.data_type = "s"

        # B: Name
        ws.cell(row=row_idx, column=2, value=name)

        # C: Promo Number (duplicate) as text
        c = ws.cell(row=row_idx, column=3, value=number)
        c.data_type = "s"

        # D: Promo Name (duplicate of Name)
        ws.cell(row=row_idx, column=4, value=name)

        # E: Cue (wrap text for readability)
        e_cell = ws.cell(row=row_idx, column=5, value=cue)
        e_cell.alignment = Alignment(wrap_text=True)

        # F: Notes (empty)
        ws.cell(row=row_idx, column=6, value="")

        # G: =LEN(E{row})
        ws.cell(row=row_idx, column=7, value=f"=LEN(E{row_idx})")

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_path)
    print(f"Saved {len(rows)} rows → {output_path}")


if __name__ == "__main__":
    import sys
    from parser import parse_doc

    in_path = sys.argv[1] if len(sys.argv) > 1 else "/Users/ryanh/Downloads/UFC 327 POC for Scripts - V2.docx"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "/tmp/test_output.xlsx"

    rows = parse_doc(in_path)
    write_xlsx(rows, out_path)
